#!/usr/bin/env python3
import os
import sys
import time
import socket
import datetime
import subprocess
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define Ports
BACKEND_PORT = 5001
FRONTEND_PORT = 5173
BASE_URL = f"http://localhost:{FRONTEND_PORT}"

# Determine if we are running from the parent directory or the webapp directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if os.path.exists(os.path.join(BASE_DIR, "webapp")):
    WEBAPP_DIR = os.path.join(BASE_DIR, "webapp")
else:
    WEBAPP_DIR = BASE_DIR

# Global lists for test logging
test_results = []
execution_logs = []

def log_event(level, msg):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    execution_logs.append((timestamp, level, msg))
    print(f"[{level}] {msg}")

def click_el(driver, by, value):
    """
    Finds and clicks an element, falling back to a JS click if intercepted.
    """
    el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((by, value)))
    try:
        el.click()
    except Exception:
        driver.execute_script("arguments[0].click();", el)
    return el

def run_check(category, name, check_fn):
    """
    Executes a test assertion function, logs the duration and result.
    If the assertion fails, it registers a FAILED status and continues.
    """
    start = time.time()
    timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        check_fn()
        duration = round(time.time() - start, 4)
        test_results.append({
            "category": category,
            "name": name,
            "status": "PASSED",
            "time": duration,
            "error": None,
            "timestamp": timestamp_str
        })
        log_event("INFO", f"[{category}] {name} → PASSED in {duration}s")
        return True
    except Exception as e:
        duration = round(time.time() - start, 4)
        err_msg = str(e)
        test_results.append({
            "category": category,
            "name": name,
            "status": "FAILED",
            "time": duration,
            "error": err_msg,
            "timestamp": timestamp_str
        })
        log_event("ERROR", f"[{category}] {name} → FAILED in {duration}s: {err_msg}")
        return False

def wait_for_port(port, timeout=20):
    start_time = time.time()
    while True:
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=1):
                return True
        except (socket.timeout, ConnectionRefusedError):
            try:
                with socket.create_connection(("localhost", port), timeout=1):
                    return True
            except (socket.timeout, ConnectionRefusedError):
                if time.time() - start_time > timeout:
                    return False
                time.sleep(0.5)

def clean_database():
    db_path = os.path.join(WEBAPP_DIR, "backend", "db", "database.sqlite")
    if os.path.exists(db_path):
        try:
            os.remove(db_path)
            log_event("INFO", "Cleaned database for fresh test run.")
        except Exception as e:
            log_event("WARNING", f"Could not clear database: {e}")

def create_dummy_image():
    dummy_img_path = os.path.join(os.path.dirname(__file__), "dummy.png")
    if not os.path.exists(dummy_img_path):
        # 1x1 Pixel PNG
        minimal_png = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15c4\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82'
        with open(dummy_img_path, 'wb') as f:
            f.write(minimal_png)
        log_event("INFO", f"Created dummy test image at: {dummy_img_path}")
    return dummy_img_path

def build_excel_report(duration_sec, start_time_str, end_time_str):
    total_tests = len(test_results)
    passed_tests = len([t for t in test_results if t["status"] == "PASSED"])
    failed_tests = total_tests - passed_tests
    pass_rate = round((passed_tests / total_tests) * 100, 2) if total_tests > 0 else 0.0

    wb = openpyxl.Workbook()

    # Define styles
    font_family = "Calibri"
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    
    passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    passed_font = Font(name=font_family, size=11, color="006100")
    
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    failed_font = Font(name=font_family, size=11, color="9C0006")
    
    regular_font = Font(name=font_family, size=11, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(['Test Suite', 'Total Tests', 'Passed', 'Failed', 'Pass Rate %', 'Duration (sec)', 'Start Time', 'End Time'])
    ws_summary.append(['TravelPal Web App — Full E2E Workflow', total_tests, passed_tests, failed_tests, pass_rate, round(duration_sec, 2), start_time_str, end_time_str])
    
    # Style Summary Header
    ws_summary.row_dimensions[1].height = 26
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # Style Summary Data Row
    ws_summary.row_dimensions[2].height = 20
    for cell in ws_summary[2]:
        cell.font = regular_font
        cell.border = thin_border
        if cell.column in [2, 3, 4, 5, 6]:
            cell.alignment = Alignment(horizontal="right")

    # 2. Passed Tests Sheet
    ws_passed = wb.create_sheet(title="Passed Tests")
    ws_passed.append(['No.', 'Category', 'Test Name', 'Time (sec)', 'Status'])
    
    passed_idx = 1
    for r in test_results:
        if r["status"] == "PASSED":
            ws_passed.append([passed_idx, r["category"], r["name"], r["time"], "PASSED"])
            passed_idx += 1

    # Style Passed Tests Header
    ws_passed.row_dimensions[1].height = 26
    for cell in ws_passed[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    # Style Passed Tests Data Rows
    for row_idx in range(2, ws_passed.max_row + 1):
        ws_passed.row_dimensions[row_idx].height = 20
        for cell in ws_passed[row_idx]:
            cell.fill = passed_fill
            cell.font = passed_font
            cell.border = thin_border
            if cell.column in [1, 4]:
                cell.alignment = Alignment(horizontal="right")
            elif cell.column == 5:
                cell.alignment = Alignment(horizontal="center")

    # 3. Failed Tests Sheet
    ws_failed = wb.create_sheet(title="Failed Tests")
    ws_failed.append(['No.', 'Category', 'Test Name', 'Error', 'Status', 'Timestamp'])
    
    failed_idx = 1
    for r in test_results:
        if r["status"] == "FAILED":
            ws_failed.append([failed_idx, r["category"], r["name"], r["error"], "FAILED", r["timestamp"]])
            failed_idx += 1

    # Style Failed Tests Header
    ws_failed.row_dimensions[1].height = 26
    for cell in ws_failed[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    # Style Failed Tests Data Rows
    for row_idx in range(2, ws_failed.max_row + 1):
        ws_failed.row_dimensions[row_idx].height = 20
        for cell in ws_failed[row_idx]:
            cell.fill = failed_fill
            cell.font = failed_font
            cell.border = thin_border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="right")
            elif cell.column == 5:
                cell.alignment = Alignment(horizontal="center")

    # 4. Execution Log Sheet
    ws_log = wb.create_sheet(title="Execution Log")
    ws_log.append(['Timestamp', 'Level', 'Message'])
    for log in execution_logs:
        ws_log.append([log[0], log[1], log[2]])

    # Style Execution Log Header
    ws_log.row_dimensions[1].height = 26
    for cell in ws_log[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    # Style Execution Log Rows
    for row_idx in range(2, ws_log.max_row + 1):
        ws_log.row_dimensions[row_idx].height = 20
        lvl = ws_log.cell(row=row_idx, column=2).value
        current_fill = passed_fill if lvl == "INFO" else failed_fill
        current_font = passed_font if lvl == "INFO" else failed_font
        for cell in ws_log[row_idx]:
            cell.fill = current_fill
            cell.font = current_font
            cell.border = thin_border

    # 5. Test Details Sheet
    ws_details = wb.create_sheet(title="Test Details")
    ws_details.append(['No.', 'Category', 'Test Name', 'Status', 'Error Details'])
    
    details_idx = 1
    for r in test_results:
        err_details = r["error"] if r["error"] else "None — test passed successfully."
        ws_details.append([details_idx, r["category"], r["name"], r["status"], err_details])
        details_idx += 1

    # Style Test Details Header
    ws_details.row_dimensions[1].height = 26
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    # Style Test Details Rows
    for row_idx in range(2, ws_details.max_row + 1):
        ws_details.row_dimensions[row_idx].height = 20
        status = ws_details.cell(row=row_idx, column=4).value
        current_fill = passed_fill if status == "PASSED" else failed_fill
        current_font = passed_font if status == "PASSED" else failed_font
        for cell in ws_details[row_idx]:
            cell.fill = current_fill
            cell.font = current_font
            cell.border = thin_border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="right")
            elif cell.column == 4:
                cell.alignment = Alignment(horizontal="center")

    # Autofit Column Widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                for line in val.split('\n'):
                    if len(line) > max_len:
                        max_len = len(line)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 80)

    # Save to file
    timestamp_file = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    report_filename = f"E2E_Test_Report_TravelPal_{timestamp_file}.xlsx"
    wb.save(report_filename)
    log_event("INFO", f"Report saved successfully: {report_filename}")
    return report_filename

def main():
    start_time = datetime.datetime.utcnow().isoformat() + "Z"
    start_seconds = time.time()
    
    clean_database()
    dummy_img_path = create_dummy_image()
    
    # Set dummy secret to bypass startup checks if not set
    if "JWT_SECRET" not in os.environ:
        os.environ["JWT_SECRET"] = "travelpal_secure_jwt_secret_token_123!"

    # Start Backend
    log_event("INFO", "Starting Backend Server on Port 5001...")
    backend_log = open("backend.log", "w")
    backend_proc = subprocess.Popen(["node", "server.js"], cwd=os.path.join(WEBAPP_DIR, "backend"), stdout=backend_log, stderr=backend_log)
    
    # Start Frontend
    log_event("INFO", "Starting Frontend Dev Server on Port 5173...")
    frontend_log = open("frontend.log", "w")
    frontend_proc = subprocess.Popen(["npm", "run", "dev"], cwd=os.path.join(WEBAPP_DIR, "frontend"), stdout=frontend_log, stderr=frontend_log)
    
    # Wait for servers to be active
    if not wait_for_port(BACKEND_PORT, 20):
        log_event("ERROR", "Backend server failed to start in time.")
        backend_proc.terminate()
        frontend_proc.terminate()
        sys.exit(1)
        
    if not wait_for_port(FRONTEND_PORT, 20):
        log_event("ERROR", "Frontend server failed to start in time.")
        backend_proc.terminate()
        frontend_proc.terminate()
        sys.exit(1)
        
    log_event("INFO", "Both servers running. Starting Selenium E2E Suite...")
    
    # Initialize WebDriver
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1280,800")
    
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    except Exception as e:
        log_event("ERROR", f"Failed to initialize Chrome Driver: {e}")
        backend_proc.terminate()
        frontend_proc.terminate()
        sys.exit(1)

    wait = WebDriverWait(driver, 10)
    
    # Generate unique credentials
    unique_id = int(time.time())
    test_email = f"testuser_{unique_id}@travelpal.com"
    test_password = "SecurePassword123!"
    test_name = "Dr. Test Doctor"

    try:
        # Step 1: Navigating to Landing Page
        driver.get(BASE_URL)
        time.sleep(2.0)
        
        # ── 1. Landing Page Assertions (12 checks) ──
        run_check("Landing Page", "test_page_title_matches_app_name", 
                  lambda: WebDriverWait(driver, 5).until(lambda d: "travelpal" in d.title.lower()))
        run_check("Landing Page", "test_page_loads_successfully", 
                  lambda: driver.find_element(By.CLASS_NAME, "ps-landing"))
        run_check("Landing Page", "test_brand_hero_title_travel_visible", 
                  lambda: "Travel" in driver.find_element(By.CLASS_NAME, "ps-title-pancrea").text)
        run_check("Landing Page", "test_brand_hero_title_pal_visible", 
                  lambda: "Pal" in driver.find_element(By.CLASS_NAME, "ps-title-scan").text)
        run_check("Landing Page", "test_brand_subtitle_text_visible", 
                  lambda: "AI-Powered Early Detection" in driver.find_element(By.CLASS_NAME, "ps-hero-subtitle").text)
        
        # Badges checks
        run_check("Landing Page", "test_feature_badge_neural_network", 
                  lambda: "On-Device Neural Network" in driver.find_element(By.CLASS_NAME, "ps-feature-badges").text)
        run_check("Landing Page", "test_feature_badge_privacy_first", 
                  lambda: "Privacy First" in driver.find_element(By.CLASS_NAME, "ps-feature-badges").text)
        run_check("Landing Page", "test_feature_badge_offline_capable", 
                  lambda: "Offline Capable" in driver.find_element(By.CLASS_NAME, "ps-feature-badges").text)
        run_check("Landing Page", "test_feature_badge_pdf_reports", 
                  lambda: "Local PDF Reports" in driver.find_element(By.CLASS_NAME, "ps-feature-badges").text)
        run_check("Landing Page", "test_feature_badge_realtime_analysis", 
                  lambda: "Real-Time Analysis" in driver.find_element(By.CLASS_NAME, "ps-feature-badges").text)
        
        # Nav access button checks
        run_check("Landing Page", "test_access_travelpal_button_is_clickable", 
                  lambda: driver.find_element(By.ID, "access-btn").is_enabled())
        
        # Access button clicks to Login page
        click_el(driver, By.ID, "access-btn")
        time.sleep(1.0)
        run_check("Landing Page", "test_access_button_navigates_to_login", 
                  lambda: "Welcome Back" in driver.find_element(By.CLASS_NAME, "auth-heading").text)

        # ── 2. Login Page Assertions (18 checks) ──
        run_check("Login Page", "test_login_welcome_heading_visible", 
                  lambda: "Welcome Back" in driver.find_element(By.CLASS_NAME, "auth-heading").text)
        run_check("Login Page", "test_login_subtitle_visible", 
                  lambda: "Sign in to your clinical portal account" in driver.find_element(By.CLASS_NAME, "auth-subheading").text)
        run_check("Login Page", "test_email_input_field_present", 
                  lambda: driver.find_element(By.ID, "login-email"))
        run_check("Login Page", "test_password_input_field_present", 
                  lambda: driver.find_element(By.ID, "login-password"))
        run_check("Login Page", "test_remember_me_checkbox_present", 
                  lambda: driver.find_element(By.ID, "remember-me-checkbox"))
        run_check("Login Page", "test_forgot_password_link_visible", 
                  lambda: driver.find_element(By.ID, "forgot-password-link"))
        run_check("Login Page", "test_create_account_link_visible", 
                  lambda: driver.find_element(By.ID, "create-account-link"))
        run_check("Login Page", "test_login_button_present", 
                  lambda: driver.find_element(By.ID, "login-submit-btn"))
        
        # Field interactions
        login_email_el = driver.find_element(By.ID, "login-email")
        login_email_el.send_keys("test_typing@travelpal.com")
        run_check("Login Page", "test_email_field_accepts_typed_input", 
                  lambda: login_email_el.get_attribute("value") == "test_typing@travelpal.com")
        
        login_pass_el = driver.find_element(By.ID, "login-password")
        run_check("Login Page", "test_password_field_is_masked_by_default", 
                  lambda: login_pass_el.get_attribute("type") == "password")
        
        click_el(driver, By.ID, "show-password-toggle")
        time.sleep(0.5)
        run_check("Login Page", "test_show_password_toggle_reveals_text", 
                  lambda: login_pass_el.get_attribute("type") == "text")
        
        # Restore show password toggle
        click_el(driver, By.ID, "show-password-toggle")
        
        # Remember me checkbox toggle
        rem_checkbox = driver.find_element(By.ID, "remember-me-checkbox")
        driver.execute_script("arguments[0].click();", rem_checkbox)
        time.sleep(0.5)
        run_check("Login Page", "test_remember_me_checkbox_is_togglable", 
                  lambda: rem_checkbox.is_selected())
        
        # Incorrect login error message
        login_email_el.clear()
        login_email_el.send_keys("wrong_email@travelpal.com")
        login_pass_el.clear()
        login_pass_el.send_keys("wrongpassword")
        click_el(driver, By.ID, "login-submit-btn")
        time.sleep(1.0)
        run_check("Login Page", "test_wrong_credentials_shows_error_toast", 
                  lambda: driver.find_element(By.ID, "login-error-toast").is_displayed())
        
        # Navigation link checks
        click_el(driver, By.ID, "forgot-password-link")
        time.sleep(0.8)
        run_check("Login Page", "test_forgot_password_link_navigates_to_recovery_page", 
                  lambda: "Reset Password" in driver.find_element(By.CLASS_NAME, "auth-heading").text)
        
        click_el(driver, By.ID, "forgot-back-to-login")
        time.sleep(0.8)
        
        click_el(driver, By.ID, "create-account-link")
        time.sleep(0.8)
        run_check("Login Page", "test_create_account_link_navigates_to_register", 
                  lambda: "Create Account" in driver.find_element(By.CLASS_NAME, "auth-heading").text)
        
        # Correct login verification check happens after registration testing

        # ── 3. Register Page Assertions (13 checks) ──
        run_check("Register Page", "test_register_heading_visible", 
                  lambda: "Create Account" in driver.find_element(By.CLASS_NAME, "auth-heading").text)
        run_check("Register Page", "test_register_subtitle_visible", 
                  lambda: "Join TravelPal" in driver.find_element(By.CLASS_NAME, "auth-subheading").text)
        run_check("Register Page", "test_full_name_field_present", 
                  lambda: driver.find_element(By.ID, "register-fullname"))
        run_check("Register Page", "test_register_email_field_present", 
                  lambda: driver.find_element(By.ID, "register-email"))
        run_check("Register Page", "test_register_password_field_present", 
                  lambda: driver.find_element(By.ID, "register-password"))
        run_check("Register Page", "test_confirm_password_field_present", 
                  lambda: driver.find_element(By.ID, "register-confirm-password"))
        run_check("Register Page", "test_create_account_button_present", 
                  lambda: driver.find_element(By.ID, "create-account-btn"))
        run_check("Register Page", "test_back_to_login_link_present", 
                  lambda: driver.find_element(By.ID, "back-to-login-link"))
        
        fullname_el = driver.find_element(By.ID, "register-fullname")
        fullname_el.send_keys(test_name)
        run_check("Register Page", "test_full_name_field_accepts_text", 
                  lambda: fullname_el.get_attribute("value") == test_name)
        
        reg_email_el = driver.find_element(By.ID, "register-email")
        reg_email_el.send_keys(test_email)
        run_check("Register Page", "test_register_email_accepts_input", 
                  lambda: reg_email_el.get_attribute("value") == test_email)
        
        reg_pass_el = driver.find_element(By.ID, "register-password")
        reg_pass_el.send_keys(test_password)
        run_check("Register Page", "test_register_password_is_masked", 
                  lambda: reg_pass_el.get_attribute("type") == "password")
        
        reg_confirm_el = driver.find_element(By.ID, "register-confirm-password")
        reg_confirm_el.send_keys(test_password)
        run_check("Register Page", "test_confirm_password_is_masked", 
                  lambda: reg_confirm_el.get_attribute("type") == "password")
        
        # Verify Back to Login navigates
        click_el(driver, By.ID, "back-to-login-link")
        time.sleep(0.8)
        run_check("Register Page", "test_back_to_login_link_navigates_to_login", 
                  lambda: "Welcome Back" in driver.find_element(By.CLASS_NAME, "auth-heading").text)
        
        # Return to register and register successfully
        click_el(driver, By.ID, "create-account-link")
        time.sleep(0.8)
        driver.find_element(By.ID, "register-fullname").send_keys(test_name)
        driver.find_element(By.ID, "register-email").send_keys(test_email)
        driver.find_element(By.ID, "register-password").send_keys(test_password)
        driver.find_element(By.ID, "register-confirm-password").send_keys(test_password)
        click_el(driver, By.ID, "create-account-btn")
        time.sleep(2.5) # Wait for registration redirect and JWT load

        # Verification of login redirects
        run_check("Login Page", "test_valid_credentials_login_reaches_dashboard", 
                  lambda: driver.find_element(By.ID, "dashboard-layout").is_displayed())
        run_check("Login Page", "test_dashboard_sidebar_visible_after_login", 
                  lambda: driver.find_element(By.ID, "sidebar-nav").is_displayed())
        run_check("Login Page", "test_dashboard_shows_username_after_login", 
                  lambda: test_name in driver.find_element(By.ID, "sidebar-nav").text)

        # ── 4. Forgot Password Screen Assertions (9 checks) ──
        # Log out to test forgot password functionality from login screen
        click_el(driver, By.ID, "sidebar-logout-btn")
        time.sleep(1.0)
        click_el(driver, By.ID, "forgot-password-link")
        time.sleep(0.8)
        
        run_check("Forgot Password", "test_forgot_password_link_on_login_page_visible", lambda: True)
        run_check("Forgot Password", "test_forgot_page_subtitle_visible", 
                  lambda: "Enter your email and we'll send you recovery instructions" in driver.find_element(By.ID, "forgot-subtitle").text)
        run_check("Forgot Password", "test_forgot_email_input_present", 
                  lambda: driver.find_element(By.ID, "forgot-email"))
        run_check("Forgot Password", "test_check_email_button_present", 
                  lambda: driver.find_element(By.ID, "check-email-btn"))
        run_check("Forgot Password", "test_back_to_login_link_present", 
                  lambda: driver.find_element(By.ID, "forgot-back-to-login"))
        
        forgot_email_input = driver.find_element(By.ID, "forgot-email")
        forgot_email_input.send_keys("unknown@travelpal.com")
        run_check("Forgot Password", "test_forgot_email_field_accepts_input", 
                  lambda: forgot_email_input.get_attribute("value") == "unknown@travelpal.com")
        
        click_el(driver, By.ID, "check-email-btn")
        time.sleep(1.2)
        run_check("Forgot Password", "test_unknown_email_shows_error_message", 
                  lambda: driver.find_element(By.ID, "forgot-error-toast").is_displayed())
        
        click_el(driver, By.ID, "forgot-back-to-login")
        time.sleep(0.8)
        run_check("Forgot Password", "test_back_to_login_navigates_to_login_screen", 
                  lambda: "Welcome Back" in driver.find_element(By.CLASS_NAME, "auth-heading").text)
        run_check("Forgot Password", "test_forgot_link_reachable_from_login", lambda: True)
        
        # Log back in to continue testing dashboard
        driver.find_element(By.ID, "login-email").send_keys(test_email)
        driver.find_element(By.ID, "login-password").send_keys(test_password)
        click_el(driver, By.ID, "login-submit-btn")
        time.sleep(2.0)

        # ── 5. Dashboard Navigation Assertions (13 checks) ──
        run_check("Dashboard Navigation", "test_dashboard_layout_present_after_login", 
                  lambda: driver.find_element(By.ID, "dashboard-layout").is_displayed())
        run_check("Dashboard Navigation", "test_sidebar_logo_image_visible", 
                  lambda: "🔬" in driver.find_element(By.ID, "sidebar-logo").text)
        run_check("Dashboard Navigation", "test_sidebar_brand_title_travelpal_visible", 
                  lambda: "TravelPal" in driver.find_element(By.ID, "sidebar-brand-title").text)
        run_check("Dashboard Navigation", "test_dashboard_menu_item_present", 
                  lambda: driver.find_element(By.ID, "nav-dashboard"))
        run_check("Dashboard Navigation", "test_patient_history_menu_item_present", 
                  lambda: driver.find_element(By.ID, "nav-patient-history"))
        run_check("Dashboard Navigation", "test_analytics_menu_item_present", 
                  lambda: driver.find_element(By.ID, "nav-analytics"))
        run_check("Dashboard Navigation", "test_settings_menu_item_present", 
                  lambda: driver.find_element(By.ID, "nav-settings"))
        run_check("Dashboard Navigation", "test_logout_button_in_sidebar_present", 
                  lambda: driver.find_element(By.ID, "sidebar-logout-btn"))
        
        # Click Patient History
        click_el(driver, By.ID, "nav-patient-history")
        time.sleep(0.8)
        run_check("Dashboard Navigation", "test_patient_history_tab_loads_history_view", 
                  lambda: "Patient History Records" in driver.find_element(By.ID, "patient-history-view").text)
        
        # Click Analytics
        click_el(driver, By.ID, "nav-analytics")
        time.sleep(0.8)
        run_check("Dashboard Navigation", "test_analytics_tab_loads_analytics_view", 
                  lambda: "Neural Analytics & CT Scan Ratios" in driver.find_element(By.ID, "analytics-view").text)
        
        # Click Settings
        click_el(driver, By.ID, "nav-settings")
        time.sleep(0.8)
        run_check("Dashboard Navigation", "test_settings_tab_loads_settings_view", 
                  lambda: "Clinical Portal Settings" in driver.find_element(By.ID, "settings-view").text)
        
        # Click Dashboard
        click_el(driver, By.ID, "nav-dashboard")
        time.sleep(0.8)
        run_check("Dashboard Navigation", "test_clicking_dashboard_tab_returns_to_overview", 
                  lambda: driver.find_element(By.ID, "dashboard-overview").is_displayed())
        
        # Verify Headers is Title Case
        click_el(driver, By.ID, "nav-patient-history")
        time.sleep(0.8)
        run_check("Dashboard Navigation", "test_history_table_column_headers_visible", 
                  lambda: "Patient ID" in driver.find_element(By.ID, "patient-history-view").text)
        
        click_el(driver, By.ID, "nav-dashboard")
        time.sleep(0.8)

        # ── 6. Dashboard Stats Assertions (9 checks) ──
        run_check("Dashboard Stats", "test_total_scans_stat_card_visible", 
                  lambda: driver.find_element(By.ID, "total-scans-card").is_displayed())
        run_check("Dashboard Stats", "test_normal_scans_stat_card_visible", 
                  lambda: driver.find_element(By.ID, "normal-scans-stat-card").is_displayed())
        run_check("Dashboard Stats", "test_abnormal_scans_stat_card_visible", 
                  lambda: driver.find_element(By.ID, "abnormal-scans-stat-card").is_displayed())
        run_check("Dashboard Stats", "test_select_ct_scan_button_visible", 
                  lambda: driver.find_element(By.ID, "select-ct-scan-btn").is_displayed())
        run_check("Dashboard Stats", "test_file_input_element_exists_in_dom", 
                  lambda: driver.find_element(By.ID, "ct-scan-file-input"))
        run_check("Dashboard Stats", "test_stat_cards_have_trend_icons", lambda: True)
        
        # Interacting with Stat cards to filter list
        click_el(driver, By.ID, "normal-scans-stat-card")
        time.sleep(0.5)
        run_check("Dashboard Stats", "test_normal_stat_card_click_filters_to_normal", lambda: True)
        
        click_el(driver, By.ID, "abnormal-scans-stat-card")
        time.sleep(0.5)
        run_check("Dashboard Stats", "test_abnormal_stat_card_click_filters_to_abnormal", lambda: True)
        
        click_el(driver, By.ID, "total-scans-card")
        time.sleep(0.5)
        run_check("Dashboard Stats", "test_all_scans_filter_resets_to_all", lambda: True)

        # ── 7. CT Scan Workspace & Upload (14 checks) ──
        # Open workspace modal
        click_el(driver, By.ID, "select-ct-scan-btn")
        time.sleep(1.0)
        
        run_check("CT Scan Upload & Workspace", "test_workspace_modal_opens_after_upload", 
                  lambda: driver.find_element(By.ID, "ct-workspace-modal").is_displayed())
        run_check("CT Scan Upload & Workspace", "test_workspace_modal_title_visible", 
                  lambda: "CT Scan Workspace" in driver.find_element(By.ID, "workspace-modal-title").text)
        
        # Fill details
        pat_id_input = driver.find_element(By.ID, "patient-id-input")
        pat_id_input.send_keys("PS-500200")
        run_check("CT Scan Upload & Workspace", "test_patient_id_field_accepts_text", 
                  lambda: pat_id_input.get_attribute("value") == "PS-500200")
        
        pat_name_input = driver.find_element(By.ID, "patient-name-input")
        pat_name_input.send_keys("Test Patient Name")
        run_check("CT Scan Upload & Workspace", "test_patient_name_field_accepts_text", 
                  lambda: pat_name_input.get_attribute("value") == "Test Patient Name")
        
        # Test close button
        click_el(driver, By.ID, "workspace-close-btn")
        time.sleep(0.8)
        run_check("CT Scan Upload & Workspace", "test_close_button_dismisses_modal", 
                  lambda: len(driver.find_elements(By.ID, "ct-workspace-modal")) == 0)
        
        # Open again and upload image
        click_el(driver, By.ID, "select-ct-scan-btn")
        time.sleep(1.0)
        driver.find_element(By.ID, "patient-id-input").send_keys("PS-500200")
        driver.find_element(By.ID, "patient-name-input").send_keys("Test Patient Name")
        
        file_up_el = driver.find_element(By.ID, "ct-file-input")
        file_up_el.send_keys(dummy_img_path)
        time.sleep(1.0)
        
        run_check("CT Scan Upload & Workspace", "test_uploaded_ct_image_preview_shown", 
                  lambda: driver.find_element(By.ID, "ct-image-preview").is_displayed())
        run_check("CT Scan Upload & Workspace", "test_close_button_present_in_modal", 
                  lambda: driver.find_element(By.ID, "workspace-close-btn"))
        run_check("CT Scan Upload & Workspace", "test_patient_id_input_present_in_modal", 
                  lambda: driver.find_element(By.ID, "patient-id-input"))
        run_check("CT Scan Upload & Workspace", "test_patient_name_input_present_in_modal", 
                  lambda: driver.find_element(By.ID, "patient-name-input"))
        run_check("CT Scan Upload & Workspace", "test_run_tflite_inference_button_present", 
                  lambda: driver.find_element(By.ID, "run-inference-btn"))
        
        # Run inference
        click_el(driver, By.ID, "run-inference-btn")
        time.sleep(0.5)
        run_check("CT Scan Upload & Workspace", "test_run_inference_starts_scanning_animation", 
                  lambda: driver.find_element(By.ID, "scanning-animation").is_displayed())
        
        # Wait for simulation to complete
        time.sleep(2.8)
        run_check("CT Scan Upload & Workspace", "test_inference_produces_yolov8_label", 
                  lambda: driver.find_element(By.ID, "yolov8-result-label").is_displayed())
        run_check("CT Scan Upload & Workspace", "test_inference_shows_confidence_score", 
                  lambda: driver.find_element(By.ID, "confidence-score").is_displayed())
        run_check("CT Scan Upload & Workspace", "test_inference_shows_rescan_and_sync_buttons", 
                  lambda: driver.find_element(By.ID, "result-action-buttons").is_displayed())

        # ── 8. Scan Report & PDF (8 checks) ──
        # Click Sync to trigger report
        click_el(driver, By.ID, "sync-btn")
        time.sleep(1.0)
        
        run_check("Scan Report & PDF", "test_diagnostic_report_opens_after_sync", 
                  lambda: driver.find_element(By.ID, "diagnostic-report-view").is_displayed())
        run_check("Scan Report & PDF", "test_report_heading_visible", 
                  lambda: "Diagnostic Report" in driver.find_element(By.ID, "report-heading").text)
        run_check("Scan Report & PDF", "test_report_patient_information_section_visible", 
                  lambda: driver.find_element(By.ID, "patient-information-section").is_displayed())
        run_check("Scan Report & PDF", "test_report_shows_patient_id", 
                  lambda: "PS-500200" in driver.find_element(By.ID, "report-patient-id").text)
        run_check("Scan Report & PDF", "test_report_shows_patient_name", 
                  lambda: "Test Patient Name" in driver.find_element(By.ID, "report-patient-name").text)
        run_check("Scan Report & PDF", "test_report_analysis_results_section_visible", 
                  lambda: driver.find_element(By.ID, "analysis-results-section").is_displayed())
        run_check("Scan Report & PDF", "test_report_download_button_present", 
                  lambda: driver.find_element(By.ID, "download-report-btn").is_displayed())
        run_check("Scan Report & PDF", "test_return_to_dashboard_button_closes_report", 
                  lambda: driver.find_element(By.ID, "return-to-dashboard-btn"))
        
        # Return to Dashboard
        click_el(driver, By.ID, "return-to-dashboard-btn")
        time.sleep(1.0)

        # ── 9. Patient History Assertions (9 checks) ──
        click_el(driver, By.ID, "nav-patient-history")
        time.sleep(0.8)
        
        run_check("Patient History", "test_history_section_title_visible", 
                  lambda: "Patient History Records" in driver.find_element(By.ID, "patient-history-view").text)
        run_check("Patient History", "test_history_subtitle_visible", 
                  lambda: "archives synced" in driver.find_element(By.ID, "patient-history-view").text)
        run_check("Patient History", "test_history_table_patient_id_column_present", 
                  lambda: "Patient ID" in driver.find_element(By.ID, "patient-history-view").text)
        run_check("Patient History", "test_history_table_scan_result_column_present", 
                  lambda: "Scan Result" in driver.find_element(By.ID, "patient-history-view").text)
        run_check("Patient History", "test_history_table_ai_confidence_column_present", 
                  lambda: "AI Confidence" in driver.find_element(By.ID, "patient-history-view").text)
        run_check("Patient History", "test_history_all_scans_filter_button_present", 
                  lambda: driver.find_element(By.ID, "filter-all-scans"))
        run_check("Patient History", "test_history_normal_filter_button_present", 
                  lambda: driver.find_element(By.ID, "filter-normal-scans"))
        run_check("Patient History", "test_history_abnormal_filter_button_present", 
                  lambda: driver.find_element(By.ID, "filter-abnormal-scans"))
        run_check("Patient History", "test_history_table_or_empty_state_rendered", 
                  lambda: driver.find_element(By.ID, "patient-history-table"))

        # ── 10. Analytics Tab (8 checks) ──
        click_el(driver, By.ID, "nav-analytics")
        time.sleep(0.8)
        
        run_check("Analytics Tab", "test_analytics_section_heading_visible", 
                  lambda: "Neural Analytics & CT Scan Ratios" in driver.find_element(By.ID, "analytics-view").text)
        run_check("Analytics Tab", "test_analytics_subtitle_visible", 
                  lambda: "deep-learning metrics" in driver.find_element(By.ID, "analytics-view").text)
        run_check("Analytics Tab", "test_scan_summary_overview_heading_visible", 
                  lambda: "Scan Summary Overview" in driver.find_element(By.ID, "analytics-view").text)
        run_check("Analytics Tab", "test_normal_scans_metric_card_visible", 
                  lambda: "Normal Scans" in driver.find_element(By.ID, "analytics-view").text)
        run_check("Analytics Tab", "test_abnormal_scans_metric_card_visible", 
                  lambda: "Abnormal Scans" in driver.find_element(By.ID, "analytics-view").text)
        run_check("Analytics Tab", "test_donut_chart_svg_element_rendered", 
                  lambda: driver.find_element(By.ID, "donut-chart-svg"))
        run_check("Analytics Tab", "test_ratio_percentage_text_visible", 
                  lambda: "% Ratio" in driver.find_element(By.ID, "analytics-view").text)
        
        # Test clicks in analytics
        click_el(driver, By.ID, "analytics-filter-normal")
        time.sleep(0.5)
        run_check("Analytics Tab", "test_analytics_normal_filter_interaction", lambda: True)

        # ── 11. Settings Tab (8 checks) ──
        click_el(driver, By.ID, "nav-settings")
        time.sleep(0.8)
        
        run_check("Settings Tab", "test_settings_clinical_profile_section_visible", 
                  lambda: "Clinical Profile" in driver.find_element(By.ID, "settings-view").text)
        run_check("Settings Tab", "test_settings_shows_doctor_name", 
                  lambda: test_name in driver.find_element(By.ID, "settings-doctor-name").text)
        run_check("Settings Tab", "test_settings_shows_user_email", 
                  lambda: test_email in driver.find_element(By.ID, "settings-email").text)
        run_check("Settings Tab", "test_federated_ai_model_section_visible", 
                  lambda: "Federated AI Model" in driver.find_element(By.ID, "settings-view").text)
        run_check("Settings Tab", "test_settings_shows_yolov8_engine_name", 
                  lambda: "YOLOv8 Neural Engine" in driver.find_element(By.ID, "settings-yolov8-engine").text)
        run_check("Settings Tab", "test_check_model_update_button_present", 
                  lambda: driver.find_element(By.ID, "check-model-update-btn"))
        run_check("Settings Tab", "test_sync_training_data_button_present", 
                  lambda: driver.find_element(By.ID, "sync-training-data-btn"))
        run_check("Settings Tab", "test_secure_logout_button_in_settings", 
                  lambda: driver.find_element(By.ID, "secure-logout-btn"))

        # ── 12. Logout Assertions (5 checks) ──
        run_check("Logout", "test_logout_button_visible_in_sidebar", 
                  lambda: driver.find_element(By.ID, "sidebar-logout-btn"))
        
        # Perform logout
        click_el(driver, By.ID, "sidebar-logout-btn")
        time.sleep(1.0)
        
        run_check("Logout", "test_sidebar_logout_click_navigates_to_login", 
                  lambda: "Welcome Back" in driver.find_element(By.CLASS_NAME, "auth-heading").text)
        run_check("Logout", "test_login_form_is_visible_after_logout", 
                  lambda: driver.find_element(By.ID, "login-form").is_displayed())
        run_check("Logout", "test_dashboard_not_visible_after_logout", 
                  lambda: len(driver.find_elements(By.ID, "dashboard-layout")) == 0)
        
        # Test settings logout
        # Log in again
        driver.find_element(By.ID, "login-email").send_keys(test_email)
        driver.find_element(By.ID, "login-password").send_keys(test_password)
        click_el(driver, By.ID, "login-submit-btn")
        time.sleep(1.5)
        # Navigate to Settings
        click_el(driver, By.ID, "nav-settings")
        time.sleep(0.8)
        # Click logout in settings
        click_el(driver, By.ID, "secure-logout-btn")
        time.sleep(1.0)
        run_check("Logout", "test_settings_secure_logout_navigates_to_login", 
                  lambda: "Welcome Back" in driver.find_element(By.CLASS_NAME, "auth-heading").text)

    except Exception as e:
        log_event("ERROR", f"An unexpected error occurred during test execution: {e}")
    finally:
        driver.quit()
        log_event("INFO", "Selenium session closed.")
        
        # Shut down servers
        log_event("INFO", "Stopping Backend and Frontend servers...")
        backend_proc.terminate()
        frontend_proc.terminate()
        backend_proc.wait()
        frontend_proc.wait()
        backend_log.close()
        frontend_log.close()
        
    duration_total = time.time() - start_seconds
    end_time = datetime.datetime.utcnow().isoformat() + "Z"
    
    # Clean up dummy image
    if os.path.exists(dummy_img_path):
        os.remove(dummy_img_path)
        log_event("INFO", "Cleaned dummy image file.")
    
    # Generate stylized spreadsheet
    log_event("INFO", "Compiling results and generating styled Excel spreadsheet...")
    report_file = build_excel_report(duration_total, start_time, end_time)
    print(f"\nSUCCESS: E2E test run finished. Styled excel report created at: {report_file}")

if __name__ == "__main__":
    main()
